from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import *
from utils import *
from yfinance.utils import get_json 
import yfinance as yf


class excelReport:

    def excelWriter(self, tickers, filename):
        wb = Workbook()

        progress = Label(self, text="Collecting historic data..",bg=buttonColor,fg=buttonTextColour)
        progress.grid(row=4,column=1, columnspan=2, sticky=W+E+N+S,padx=10, pady=10)

        stockData = yf.download(tickers, period='10y')['Adj Close']
        stockData = stockData[tickers]
        wb.create_sheet(title='Stock Data')
        stockSheet = wb['Stock Data']

        for row in dataframe_to_rows(stockData, index=True, header=True):
            stockSheet.append(row)
        stockSheet.column_dimensions['A'].width = len(str(stockData.index[0]))
        stockSheet.sheet_view.showGridLines = False

        minCol,minRow,maxCol = 1, 3, 1

        for ticker in tickers:
            progress = Label(self, text="Creating report for " + ticker,bg=buttonColor,fg=buttonTextColour)
            progress.grid(row=4,column=1, columnspan=2, sticky=W+E+N+S,padx=10, pady=10)

            wb.create_sheet(title=ticker)
            sheet = wb[ticker]
            sheet.sheet_view.showGridLines = False

            try:
                excelReport.collectData(self,ticker)
            except (KeyError, TypeError):
                sheet['B2'].value = "No data available"
                sheet['B2'].font = Font(italic=True)
                continue

            sheet['B2'].value = self.tickerName
            sheet['B2'].font = Font(bold=True, size=15)
            sheet['B2'].alignment = Alignment(horizontal='left')
            sheet.merge_cells('B2:M2')

            sheet['B3'].value = self.businessSummary
            sheet['B3'].alignment = Alignment(wrap_text=True,vertical='center', horizontal='left')
            sheet.merge_cells('B3:M3')
            sheet.row_dimensions[3].height = 100

            sheet['B4'].value = "Sector Holdings"
            sheet['B4'].font = Font(bold=True)

            sheet['B17'].value = "Top Company Holdings"
            sheet['B17'].font = Font(bold=True)

            sheet['E19'].value = "Risk Statistics"
            sheet['E19'].font = Font(bold=True)
            sheet['E19'].alignment = Alignment(horizontal='center')
            sheet.merge_cells('E19:J19')

            sheet['L21'].value = "Last Five Annual Returns"
            sheet['L21'].font = Font(bold=True)

            sheet['L4'].value = 'Key Characteristics'
            sheet['L4'].font = Font(bold=True)

            dataPlacer(self.sectorHoldings,sheet,5,2, 'B','C')
            dataPlacer(self.companyHoldings,sheet,18,2,'B','C',
                        changeKeyDimensions=False)
            dataPlacer(self.annualReturns,sheet,22,12,'L','M',
                        changeKeyDimensions=False, changeValueDimensions=False)
            dataPlacer(self.keyCharacteristics,sheet,5,12,'L','M',
                        horizonalAlignmentValue='left')

            try:
                dataPlacer(self.riskData['3y'],sheet,20,5,'E','F',False,'right',True,False)
            except KeyError:
                riskData = emptyRiskStatistics
                riskData['year'] = '3y'
                dataPlacer(riskData,sheet,20,5,'E','F',False,'right',True,False)
            try:
                dataPlacer(self.riskData['5y'],sheet,20,7,'G','H',False,'right',True,False)
            except KeyError:
                riskData = emptyRiskStatistics
                riskData['year'] = '5y'
                dataPlacer(riskData,sheet,20,7,'G','H',False,'right',True,False)
            try:
                dataPlacer(self.riskData['10y'],sheet,20,9,'I','J',False,'right',True,False)
            except KeyError:
                riskData = emptyRiskStatistics
                riskData['year'] = '10y'
                dataPlacer(riskData,sheet,20,9,'I','J',False,'right',True,False)
    
            imagePlacer(self.imageURL, sheet,'L12')
            graphPlacer(ticker,stockSheet,stockData,sheet,minCol,minRow,maxCol)
            minCol += 1
            maxCol += 1
        
        stockSheet.sheet_state = 'hidden'
        wb.remove_sheet(wb['Sheet'])
        wb.save(filename)

    def collectData(self, ticker):
        url = "https://finance.yahoo.com/quote/" + ticker
        data = get_json(url)

        fundPerformance = data['fundPerformance']
        topHoldings = data['topHoldings']
        defaultKeyStatistics = data['defaultKeyStatistics']
        summaryDetail = data['summaryDetail']

        self.tickerName = data['quoteType']['longName']
        self.businessSummary = data['assetProfile']['longBusinessSummary']

        sectorData = topHoldings['sectorWeightings']
        self.sectorHoldings = {}

        for sector in sectorData:
            for key, value in sector.items():
                self.sectorHoldings[key] = str(round(value * 100, 2)) + '%'

        companyData = topHoldings['holdings']
        self.companyHoldings = {}

        for company in companyData:
            self.companyHoldings[company['holdingName']] = str(round(company['holdingPercent'] * 100, 2)) + '%'

        annualReturnsData = fundPerformance['annualTotalReturns']['returns'][:6]
        self.annualReturns = {}

        for returns in annualReturnsData:
            if returns['annualValue'] == None:
                self.annualReturns[returns['year']] = "N/A"
            else:
                self.annualReturns[returns['year']] = str(round(returns['annualValue'] * 100, 2)) + '%'

        riskStatistics = fundPerformance['riskOverviewStatistics']['riskStatistics']
        self.riskData = {}

        for risk in riskStatistics:
            self.riskData[risk['year']] = risk

        self.imageURL = data['fundProfile']['styleBoxUrl']

        self.keyCharacteristics = {}

        for option in defaultKeyStatisticsChoices:
            if option == 'fundInceptionDate':
                self.keyCharacteristics[option] = defaultKeyStatistics[option]
                self.keyCharacteristics[option] = datetime.fromtimestamp(self.keyCharacteristics[option]).strftime('%Y-%m-%d')
            else:
                self.keyCharacteristics[option] = defaultKeyStatistics[option]

        for option in defaultsummaryDetailChoices:
            self.keyCharacteristics[option] = summaryDetail[option]



