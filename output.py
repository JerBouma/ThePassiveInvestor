def excelOutput(data):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl import Workbook

    for symbol, variable in data:
        filename = symbol + '.xlsx'
        sheetName = variable

        wb = Workbook()
        wb.save(filename)

        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        for variable in data:
            pd.DataFrame(data[variable]).to_excel(writer, sheetName)

# example = {}
# example['ABC','BCD'] = [5,4,3,2,1,]
# excelOutput(example)
    
