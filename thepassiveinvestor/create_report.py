import pandas as pd
from yahooquery import Ticker
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .collect_data import collect_data
from .config import (
    DEFAULT_KEY_STATISTICS_CHOICES,
    DEFAULT_SUMMARY_DETAIL_CHOICES,
    SECTOR_CATEGORY_MAPPING,
    EMPTY_RISK_STATISTICS,
    RISK_STATISTICS_CATEGORY_MAPPING,
)
from .utils import data_placer, image_placer, graph_placer


def create_ETF_report(tickers, filename, folder=None):
    """
    Description
    ----
    This function creates an Excel Report based on the tickers entered in the function. This function
    only accepts ETFs and can not handle any other financial product. If other tickers are included,
    by default no data is returned for that ticker.

    Input
    ----
    tickers (list or string)
        A list of tickers or a single ticker from an ETF (i.e. VOO)
    filename (string)
        The name and location of the file you wish to save the data to.
    folder (string, default is None)
        If prefered, you can seperate filename and folder.

    Output
    ----
    Returns an Excel file with the given filename with data on each ticker.
    """
    workbook = Workbook()
    stock_data = Ticker(tickers, asynchronous=True).history(period="10y")["adjclose"]
    stock_data = stock_data.unstack(level=0)
    
    stock_data.index = pd.to_datetime(stock_data.index)
    stock_data.index = stock_data.index.tz_localize(None)

    if isinstance(tickers, str):
        tickers = [tickers]

    if isinstance(stock_data, pd.DataFrame):
        stock_data = stock_data[tickers]
    elif isinstance(stock_data, pd.Series):
        stock_data = pd.DataFrame(stock_data)

    if filename[-4:] not in ["xlsx", "xlsm", "xlsb"]:
        filename = f"{filename}.xlsx"
    if folder is not None:
        filename = os.path.join(folder, filename)

    workbook.create_sheet(title="Stock Data")
    stock_sheet = workbook["Stock Data"]

    for row in dataframe_to_rows(stock_data, index=True, header=True):
        stock_sheet.append(row)
    stock_sheet.column_dimensions["A"].width = len(str(stock_data.index[0]))
    stock_sheet.sheet_view.showGridLines = False

    min_col, min_row, max_col = 1, 3, 1

    for ticker in tickers:
        workbook.create_sheet(title=ticker)
        sheet = workbook[ticker]
        sheet.sheet_view.showGridLines = False

        try:
            ticker_data = collect_data(ticker)
        except (KeyError, TypeError):
            sheet["B2"].value = "No data available"
            sheet["B2"].font = Font(italic=True)
            continue

        sheet["B2"].value = ticker_data["long_name"]
        sheet["B2"].font = Font(bold=True, size=15)
        sheet["B2"].alignment = Alignment(horizontal="left")
        sheet.merge_cells("B2:M2")

        sheet["B3"].value = ticker_data["summary"]
        sheet["B3"].alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="left"
        )
        sheet.merge_cells("B3:M3")
        sheet.row_dimensions[3].height = 100

        sheet["B4"].value = "Sector Holdings"
        sheet["B4"].font = Font(bold=True)

        sheet["B17"].value = "Top Company Holdings"
        sheet["B17"].font = Font(bold=True)

        sheet["E19"].value = "Risk Statistics"
        sheet["E19"].font = Font(bold=True)
        sheet["E19"].alignment = Alignment(horizontal="center")
        sheet.merge_cells("E19:J19")

        sheet["L21"].value = "Last Five Annual Returns"
        sheet["L21"].font = Font(bold=True)

        sheet["L4"].value = "Key Characteristics"
        sheet["L4"].font = Font(bold=True)

        for sector in ticker_data["sector_holdings"]:
            if sector in SECTOR_CATEGORY_MAPPING:
                new_sector = SECTOR_CATEGORY_MAPPING[sector]
                ticker_data["sector_holdings"][new_sector] = ticker_data[
                    "sector_holdings"
                ].pop(sector)

        data_placer(
            ticker_data["sector_holdings"],
            sheet,
            5,
            2,
            "B",
            "C",
            value_percentage=True,
        )
        data_placer(
            ticker_data["company_holdings"],
            sheet,
            18,
            2,
            "B",
            "C",
            change_key_dimensions=False,
            value_percentage=True,
        )
        data_placer(
            ticker_data["annual_returns"],
            sheet,
            22,
            12,
            "L",
            "M",
            change_key_dimensions=False,
            change_value_dimensions=False,
            key_number=True,
            value_percentage=True,
        )

        for key in list(ticker_data["key_characteristics"].keys()):
            if key in DEFAULT_KEY_STATISTICS_CHOICES:
                new_key = DEFAULT_KEY_STATISTICS_CHOICES[key]
                ticker_data["key_characteristics"][new_key] = ticker_data[
                    "key_characteristics"
                ].pop(key)
            elif key in DEFAULT_SUMMARY_DETAIL_CHOICES:
                new_key = DEFAULT_SUMMARY_DETAIL_CHOICES[key]
                ticker_data["key_characteristics"][new_key] = ticker_data[
                    "key_characteristics"
                ].pop(key)

        data_placer(
            ticker_data["key_characteristics"],
            sheet,
            5,
            12,
            "L",
            "M",
            horizontal_alignment_value="left",
        )

        try:
            for risk_metric in list(ticker_data["risk_data"]["3y"].keys()):
                if risk_metric in RISK_STATISTICS_CATEGORY_MAPPING:
                    new_risk_metric = RISK_STATISTICS_CATEGORY_MAPPING[risk_metric]
                    ticker_data["risk_data"]["3y"][new_risk_metric] = ticker_data[
                        "risk_data"
                    ]["3y"].pop(risk_metric)

            data_placer(
                ticker_data["risk_data"]["3y"],
                sheet,
                20,
                5,
                "E",
                "F",
                False,
                "right",
                True,
                False,
            )
        except KeyError:
            risk_data = EMPTY_RISK_STATISTICS
            risk_data["year"] = "3y"
            data_placer(
                ticker_data["risk_data"],
                sheet,
                20,
                5,
                "E",
                "F",
                False,
                "right",
                True,
                False,
            )
        try:
            for risk_metric in list(ticker_data["risk_data"]["5y"].keys()):
                if risk_metric in RISK_STATISTICS_CATEGORY_MAPPING:
                    new_risk_metric = RISK_STATISTICS_CATEGORY_MAPPING[risk_metric]
                    ticker_data["risk_data"]["5y"][new_risk_metric] = ticker_data[
                        "risk_data"
                    ]["5y"].pop(risk_metric)

            data_placer(
                ticker_data["risk_data"]["5y"],
                sheet,
                20,
                7,
                "G",
                "H",
                False,
                "right",
                True,
                False,
            )
        except KeyError:
            risk_data = EMPTY_RISK_STATISTICS
            risk_data["year"] = "5y"
            data_placer(
                ticker_data["risk_data"],
                sheet,
                20,
                7,
                "G",
                "H",
                False,
                "right",
                True,
                False,
            )
        try:
            for risk_metric in list(ticker_data["risk_data"]["10y"].keys()):
                if risk_metric in RISK_STATISTICS_CATEGORY_MAPPING:
                    new_risk_metric = RISK_STATISTICS_CATEGORY_MAPPING[risk_metric]
                    ticker_data["risk_data"]["10y"][new_risk_metric] = ticker_data[
                        "risk_data"
                    ]["10y"].pop(risk_metric)

            data_placer(
                ticker_data["risk_data"]["10y"],
                sheet,
                20,
                9,
                "I",
                "J",
                False,
                "right",
                True,
                False,
            )
        except KeyError:
            risk_data = EMPTY_RISK_STATISTICS
            risk_data["year"] = "10y"
            data_placer(
                ticker_data["risk_data"],
                sheet,
                20,
                9,
                "I",
                "J",
                False,
                "right",
                True,
                False,
            )

        image_placer(ticker_data["image_URL"], sheet, "L12")
        graph_placer(stock_sheet, stock_data, sheet, min_col, min_row, max_col, "E4")
        min_col += 1
        max_col += 1

    try:
        workbook.remove(workbook["Sheet"])
    except KeyError:
        pass

    stock_sheet.sheet_state = "hidden"
    workbook.save(filename)
